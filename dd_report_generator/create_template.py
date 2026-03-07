"""Excel入力テンプレートを生成するスクリプト。

DDレポートに必要なデータを入力するためのExcelテンプレートを作成する。
各シートに対応するデータを入力した後、generate_report.py で PowerPoint を生成する。
"""

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER


def create_cover_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "表紙"
    headers = ["項目", "内容"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, 2)

    items = ["案件名", "対象会社名", "報告日", "作成者", "作成会社"]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 1, 2)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def create_company_overview_sheet(wb: Workbook):
    ws = wb.create_sheet("会社概要")
    headers = ["項目", "内容"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, 2)

    items = [
        "会社名", "代表者", "設立年月日", "資本金", "本社所在地",
        "従業員数", "事業内容", "主要取引先", "主要仕入先",
        "関連会社", "株主構成", "沿革・備考",
    ]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 1, 2)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60


def create_bs_sheet(wb: Workbook):
    ws = wb.create_sheet("貸借対照表")
    headers = ["勘定科目", "第1期（金額）", "第2期（金額）", "第3期（金額）", "備考"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    # 期情報入力行
    ws.cell(row=2, column=1, value="決算期").font = Font(bold=True, color="FF0000")
    ws.cell(row=2, column=2, value="例: 2022年3月期")
    ws.cell(row=2, column=3, value="例: 2023年3月期")
    ws.cell(row=2, column=4, value="例: 2024年3月期")

    assets = [
        "【資産の部】",
        "現金及び預金", "受取手形", "売掛金", "有価証券", "棚卸資産",
        "前払費用", "その他流動資産", "流動資産合計",
        "建物及び構築物", "機械装置", "土地", "建設仮勘定",
        "無形固定資産", "投資有価証券", "関係会社株式",
        "長期貸付金", "その他固定資産", "固定資産合計",
        "資産合計",
        "",
        "【負債の部】",
        "支払手形", "買掛金", "短期借入金", "未払金", "未払費用",
        "未払法人税等", "前受金", "賞与引当金",
        "その他流動負債", "流動負債合計",
        "長期借入金", "退職給付引当金", "その他固定負債", "固定負債合計",
        "負債合計",
        "",
        "【純資産の部】",
        "資本金", "資本剰余金", "利益剰余金",
        "自己株式", "その他", "純資産合計",
        "",
        "負債・純資産合計",
    ]
    for i, item in enumerate(assets, 3):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(assets) + 2, len(headers))
    ws.column_dimensions["A"].width = 25
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 30


def create_pl_sheet(wb: Workbook):
    ws = wb.create_sheet("損益計算書")
    headers = ["勘定科目", "第1期（金額）", "第2期（金額）", "第3期（金額）", "備考"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    ws.cell(row=2, column=1, value="決算期").font = Font(bold=True, color="FF0000")
    ws.cell(row=2, column=2, value="例: 2022年3月期")
    ws.cell(row=2, column=3, value="例: 2023年3月期")
    ws.cell(row=2, column=4, value="例: 2024年3月期")

    items = [
        "売上高", "売上原価", "売上総利益",
        "",
        "販売費及び一般管理費",
        "  人件費", "  地代家賃", "  減価償却費", "  その他販管費",
        "販管費合計",
        "営業利益",
        "",
        "営業外収益",
        "  受取利息", "  受取配当金", "  その他営業外収益",
        "営業外費用",
        "  支払利息", "  その他営業外費用",
        "経常利益",
        "",
        "特別利益", "特別損失",
        "税引前当期純利益",
        "法人税等", "当期純利益",
        "",
        "EBITDA（参考）",
    ]
    for i, item in enumerate(items, 3):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 2, len(headers))
    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 30


def create_adjusted_bs_sheet(wb: Workbook):
    ws = wb.create_sheet("修正貸借対照表")
    headers = [
        "勘定科目", "帳簿価額", "時価評価額", "修正額", "修正理由",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    items = [
        "【資産の部】",
        "現金及び預金", "売掛金", "棚卸資産", "有価証券",
        "土地", "建物", "投資有価証券", "関係会社株式",
        "その他資産", "資産合計（時価）",
        "",
        "【負債の部】",
        "借入金", "退職給付引当金", "簿外債務",
        "その他負債", "負債合計（時価）",
        "",
        "時価純資産",
        "過剰資産（遊休資産等）",
        "実態純資産",
    ]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 1, len(headers))
    ws.column_dimensions["A"].width = 25
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 40


def create_account_verification_sheet(wb: Workbook):
    ws = wb.create_sheet("科目別検証")
    headers = [
        "検証No.", "対象科目", "帳簿価額", "検証結果",
        "修正額", "修正後金額", "リスク評価", "詳細コメント",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    # 20行分のデータ領域を用意
    for r in range(2, 22):
        ws.cell(row=r, column=1, value=r - 1)
    _style_data_area(ws, 2, 21, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 50


def create_tax_dd_sheet(wb: Workbook):
    ws = wb.create_sheet("税務DD")
    headers = ["検証項目", "検証内容", "リスク評価", "指摘事項", "推定影響額", "対応策"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    items = [
        "法人税申告内容", "消費税処理", "源泉徴収",
        "移転価格税制", "税務調査履歴", "繰越欠損金",
        "グループ通算制度", "税効果会計", "その他税務リスク",
    ]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 1, len(headers))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 35


def create_summary_sheet(wb: Workbook):
    ws = wb.create_sheet("総括")
    headers = ["区分", "内容"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, 2)

    items = [
        "財務DD総括",
        "主要なリスク事項",
        "時価純資産の評価",
        "収益力の評価",
        "税務DD総括",
        "推定簿外債務",
        "その他留意事項",
        "総合所見",
    ]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
    _style_data_area(ws, 2, len(items) + 1, 2)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def create_template(output_path: str):
    wb = Workbook()
    create_cover_sheet(wb)
    create_company_overview_sheet(wb)
    create_bs_sheet(wb)
    create_pl_sheet(wb)
    create_adjusted_bs_sheet(wb)
    create_account_verification_sheet(wb)
    create_tax_dd_sheet(wb)
    create_summary_sheet(wb)
    wb.save(output_path)
    print(f"テンプレートを作成しました: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="DD報告書用Excelテンプレートを生成")
    parser.add_argument(
        "-o", "--output",
        default="templates/dd_input_template.xlsx",
        help="出力ファイルパス (default: templates/dd_input_template.xlsx)",
    )
    args = parser.parse_args()
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    create_template(args.output)


if __name__ == "__main__":
    main()
